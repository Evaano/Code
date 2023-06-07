import openpyxl

# Source xl file
source_workbook =openpyxl.load_workbook("source.xlsx")
source_sheet = source_workbook.active

# Destination xl file
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active

# Iterate over the rows in source
for row in source_sheet.iter_rows(values_only=True):
    for cell_value in row:
        if cell_value is not None and "Vinares" in str(cell_value):
            content = cell_value.split("Vinares", 1)[1].strip()
            content = "Vinares " + content
            destination_sheet.append([content])

# Save the destination workbook with the copied content
destination_workbook.save("destination.xlsx")
print("Content copied")

# Close workbooks
source_workbook.close()
destination_workbook.close()